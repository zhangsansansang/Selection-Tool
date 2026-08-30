# -*- coding: utf-8 -*-
"""
从 表格文档/3D轮廓传感器视野选型换算表V1.2.xlsx 的「信息页」提取换算参数 -> data/fov.js
- 排除 H 后缀型号 (退市)
- 换算公式(查询页 r6/r10/r14/r18 反推):
    视场 = 近视野 + (架设距离-最近架设)/(最远架设-最近架设) * (远视野-近视野)
    X分辨率 = 视场 / 单轮廓点数 * 1000 (μm)
    Z分辨率 = 最小Z + (视场-最小视野)/(最大视野-最小视野) * (最大Z-最小Z)
"""
import openpyxl, json, re, os

XLSX = '表格文档/3D轮廓传感器视野选型换算表V1.2.xlsx'
OUT = 'data/fov.js'

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['信息页']

# 列索引(1-based) 见 信息页 r2 表头
models = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name or not isinstance(name, str):
        continue
    if re.search(r'\dH', name):  # H 后缀退市
        continue
    def g(c):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else None
    m = {
        'model': name,
        'points': g(2),        # 单轮廓点数
        'cd': g(3),            # 推荐架设距离
        'near': g(4), 'mid': g(5), 'far': g(6),  # 近/中/远视野
        'zRange': g(7),        # 测量范围
        'minH': g(16), 'maxH': g(17),  # 最近/最远架设距离(=最小/最大高度)
        'minView': g(18), 'maxView': g(19),  # 最小/最大视野
        'minX': g(20), 'maxX': g(21),        # 最小/最大X精度
        'minZ': g(22), 'maxZ': g(23),        # 最小/最大Z精度
    }
    models.append(m)

print(f'视野换算型号数(排除H): {len(models)}')
os.makedirs('data', exist_ok=True)
js = 'window.FOV_DB=' + json.dumps(models, ensure_ascii=False) + ';\n'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(js)
print(f'已生成 {OUT} ({os.path.getsize(OUT)} bytes)')
