# -*- coding: utf-8 -*-
"""
从 表格文档/3D轮廓传感器视野帧率表.xlsx 提取帧率数据 -> data/frame_rate.js
- 排除 DP2000 V1.0 / DP3000 V1.0 (退市)
- 每 sheet 内: 帧率列(静态) = 系列级网格; 高度/宽度范围由 VLOOKUP 公式 = 测量范围/中视野 × ROI/总ROI
"""
import openpyxl, json, re, os

XLSX = '表格文档/3D轮廓传感器视野帧率表.xlsx'
OUT = 'data/frame_rate.js'

# sheet -> series key (排除 V1.0)
SHEET_SERIES = {
    'DP2000 V2.0视野帧率表': 'DP2000V2',
    'DP3000 V2.0视野帧率表': 'DP3000V2',
    'DP4000 视野帧率表': 'DP4000',
}

def is_current(name):
    if 'V1.0' in name: return False        # 退市 V1.0
    if re.search(r'\dH', name): return False  # H 后缀退市 (01H / 3H)
    return True

wb = openpyxl.load_workbook(XLSX, data_only=True)

# 1) 系列帧率网格 (静态列 A=ROI高, C=ROI宽, F=帧率)
series_data = {}
for sn, key in SHEET_SERIES.items():
    ws = wb[sn]
    # 总 ROI = MAX(A:A), MAX(C:C)
    totalH = totalW = 0
    rows = []
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        c = ws.cell(r, 3).value
        f = ws.cell(r, 6).value
        if not isinstance(a, (int, float)) or not isinstance(c, (int, float)):
            continue
        a = int(a); c = int(c)
        if isinstance(f, (int, float)):
            totalH = max(totalH, a)
            totalW = max(totalW, c)
            rows.append([a, c, round(float(f), 3)])
    series_data[key] = {'totalH': totalH, 'totalW': totalW, 'rows': rows}
    print(f'[{key}] totalH={totalH} totalW={totalW} 行数={len(rows)}')

# 2) 相机信息 -> 在售型号 (排除 H + V1.0)
ws = wb['相机信息']
models = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name or not isinstance(name, str):
        continue
    if not is_current(name):
        continue
    points = ws.cell(r, 2).value
    cd = ws.cell(r, 3).value          # ①推荐架设距离
    near = ws.cell(r, 4).value        # ②近视野
    mid = ws.cell(r, 5).value         # ③中视野
    far = ws.cell(r, 6).value         # ④远视野
    zr = ws.cell(r, 7).value          # ⑤测量范围
    # 系列判定
    if name.startswith('MV-DP2'):
        series = 'DP2000V2'
    elif name.startswith('MV-DP3'):
        series = 'DP3000V2'
    elif name.startswith('MV-DP4'):
        series = 'DP4000'
    else:
        continue
    models.append({
        'model': name, 'series': series, 'points': points,
        'cd': cd, 'near': near, 'mid': mid, 'far': far, 'zRange': zr,
    })
print(f'在售型号数: {len(models)}')

# 输出 JS
os.makedirs('data', exist_ok=True)
js = 'window.FRAME_DATA=' + json.dumps({'series': series_data, 'models': models}, ensure_ascii=False) + ';\n'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(js)
print(f'已生成 {OUT} ({os.path.getsize(OUT)} bytes)')
